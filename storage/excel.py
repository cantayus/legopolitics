from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from legopolitics.constants import DISCLAIMER, EXCEL_SHEETS
from legopolitics.storage.atomic import atomic_build
from legopolitics.storage.sqlite import flatten_record

DANGEROUS_PREFIXES = ("=", "+", "-", "@")


def sanitize_cell(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(DANGEROUS_PREFIXES):
        return "'" + value
    return value


def _frame(records: list[Any]) -> pd.DataFrame:
    rows = []
    for record in records:
        row = {k: sanitize_cell(v) for k, v in flatten_record(record).items()}
        rows.append(row)
    return pd.DataFrame(rows)


def write_workbook(
    path: Path, tables: dict[str, list[Any]], metadata: dict[str, Any], include_charts: bool = True
) -> Path:
    def build(temp: Path) -> Path:
        with pd.ExcelWriter(temp, engine="openpyxl") as writer:
            readme = pd.DataFrame(
                {
                    "Item": ["Project", "Purpose", "Trademark notice", "Authoritative formats"],
                    "Value": [
                        "legopolitics",
                        "Multimodal analysis of political brick-built videos",
                        DISCLAIMER,
                        "SQLite, Parquet, and JSONL; Excel is a presentation layer",
                    ],
                }
            )
            readme.to_excel(writer, sheet_name="README", index=False)
            for sheet in EXCEL_SHEETS:
                if sheet == "README":
                    continue
                if sheet == "Run_Metadata":
                    frame = pd.DataFrame([metadata])
                elif sheet == "Data_Dictionary":
                    frame = _data_dictionary(tables)
                else:
                    frame = _frame(tables.get(sheet, []))
                if frame.empty:
                    frame = pd.DataFrame({"status": ["No records"]})
                frame.to_excel(writer, sheet_name=sheet, index=False)
        workbook = load_workbook(temp)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            sheet.row_dimensions[1].height = 24
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(wrap_text=True)
            for column in sheet.columns:
                letter = get_column_letter(column[0].column)
                maximum = max(
                    (len(str(c.value)) if c.value is not None else 0 for c in column), default=0
                )
                sheet.column_dimensions[letter].width = min(max(10, maximum + 2), 60)
                for cell in column[1:]:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                header = str(column[0].value or "").lower()
                if header.endswith("_path") or header in {
                    "path",
                    "image_path",
                    "crop_path",
                    "report",
                }:
                    for cell in column[1:]:
                        if (
                            cell.value
                            and isinstance(cell.value, str)
                            and os.path.exists(cell.value)
                        ):
                            cell.hyperlink = Path(cell.value).resolve().as_uri()
                            cell.style = "Hyperlink"
            confidence_cols = [c.column for c in sheet[1] if "confidence" in str(c.value).lower()]
            for index in confidence_cols:
                letter = get_column_letter(index)
                sheet.conditional_formatting.add(
                    f"{letter}2:{letter}{max(2, sheet.max_row)}",
                    ColorScaleRule(
                        start_type="num",
                        start_value=0,
                        start_color="F8696B",
                        mid_type="num",
                        mid_value=0.5,
                        mid_color="FFEB84",
                        end_type="num",
                        end_value=1,
                        end_color="63BE7B",
                    ),
                )
        if include_charts and "Detections" in workbook.sheetnames:
            sheet = workbook["Detections"]
            headers = {cell.value: i + 1 for i, cell in enumerate(sheet[1])}
            if "confidence" in headers and sheet.max_row > 1:
                chart = BarChart()
                data = Reference(
                    sheet, min_col=headers["confidence"], min_row=1, max_row=min(sheet.max_row, 21)
                )
                chart.add_data(data, titles_from_data=True)
                chart.title = "Detection confidence sample"
                sheet.add_chart(chart, "AA2")
        workbook.save(temp)
        load_workbook(temp, read_only=True).close()
        return temp

    atomic_build(path, build)
    return path


def _data_dictionary(tables: dict[str, list[Any]]) -> pd.DataFrame:
    rows = []
    for sheet, records in tables.items():
        if not records:
            continue
        data = flatten_record(records[0])
        for name, value in data.items():
            rows.append(
                {
                    "sheet": sheet,
                    "field": name,
                    "example_type": type(value).__name__,
                    "description": "See package schemas and documentation.",
                }
            )
    return pd.DataFrame(rows)
