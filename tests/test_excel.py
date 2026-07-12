from pathlib import Path

from openpyxl import load_workbook

from legopolitics.constants import DISCLAIMER
from legopolitics.storage.excel import sanitize_cell, write_workbook


def test_formula_sanitization():
    assert sanitize_cell("=2+2").startswith("'")
    assert sanitize_cell("safe") == "safe"


def test_workbook_reopens_and_has_disclaimer(tmp_path: Path):
    path = write_workbook(tmp_path / "result.xlsx", {"Frames": []}, {"version": "x"}, False)
    workbook = load_workbook(path, read_only=True)
    values = [cell.value for row in workbook["README"].iter_rows() for cell in row]
    workbook.close()
    assert DISCLAIMER in values
